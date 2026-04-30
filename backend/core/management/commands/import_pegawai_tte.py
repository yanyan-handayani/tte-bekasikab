import csv
import hashlib
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from django.contrib.auth import get_user_model

from core.crypto import hash_lookup
from core.models import MstDataPegawai, RefInstansi, MstJabatan


IMPORT_USER = "pegawai_import"


def clean_text(value):
    return " ".join(str(value or "").replace("\ufeff", "").strip().split())


def clean_nip(value):
    value = clean_text(value)
    value = value.replace("'", "").replace('"', "")
    return "".join(ch for ch in value if ch.isdigit())


def normalize_text(value):
    return clean_text(value).upper()


def make_generated_username(nip):
    """
    Sesuaikan dengan helper username existing aplikasi kalau sudah ada.
    Ini fallback stabil: NIP sama = username sama.
    """
    digest = hashlib.sha256(nip.encode("utf-8")).hexdigest()
    return f"usr_{digest[:16]}"


class Command(BaseCommand):
    help = "Import/update data pegawai TTE dari CSV/XLSX + create/link user."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Path file CSV/XLSX")
        parser.add_argument("--commit", action="store_true", help="Commit ke database")
        parser.add_argument("--create-user", action="store_true", help="Create/link user")
        parser.add_argument("--delimiter", default=",", help="Delimiter CSV, default koma")
        parser.add_argument("--report", default=None, help="Path output report CSV")

    def handle(self, *args, **options):
        file_path = Path(options["file"])
        do_commit = options["commit"]
        create_user = options["create_user"]
        delimiter = options["delimiter"]

        if not file_path.exists():
            raise CommandError(f"File tidak ditemukan: {file_path}")

        report_path = options["report"] or f"report_import_pegawai_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"

        rows = self.read_input(file_path, delimiter)
        if not rows:
            raise CommandError("File kosong atau header tidak terbaca.")

        report = self.validate_rows(rows)
        self.write_report(report, report_path)

        total = len(report)
        invalid = [r for r in report if r["status"] != "OK"]

        self.stdout.write(self.style.WARNING(f"Total row     : {total}"))
        self.stdout.write(self.style.SUCCESS(f"Valid         : {total - len(invalid)}"))
        self.stdout.write(self.style.ERROR(f"Error/Skip    : {len(invalid)}"))
        self.stdout.write(self.style.WARNING(f"Report        : {report_path}"))

        if invalid:
            self.stdout.write(self.style.ERROR("Masih ada error. Perbaiki data/report dulu sebelum commit."))
            return

        if not do_commit:
            self.stdout.write(self.style.SUCCESS("DRY-RUN selesai. Belum ada perubahan database."))
            self.stdout.write("Untuk eksekusi gunakan tambahan: --commit")
            return

        with transaction.atomic():
            self.commit_rows(report, create_user=create_user)

        self.stdout.write(self.style.SUCCESS("Import/update pegawai selesai."))

    def read_input(self, file_path, delimiter):
        suffix = file_path.suffix.lower()

        if suffix == ".csv":
            return self.read_csv(file_path, delimiter)

        if suffix in [".xlsx", ".xlsm"]:
            return self.read_xlsx(file_path)

        raise CommandError("Format file harus .csv atau .xlsx")

    def read_csv(self, file_path, delimiter):
        with open(file_path, newline="", encoding="utf-8-sig") as f:
            return list(csv.DictReader(f, delimiter=delimiter))

    def read_xlsx(self, file_path):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise CommandError("openpyxl belum terinstall. Install: pip install openpyxl")

        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [clean_text(h) for h in rows[0]]
        data = []

        for row in rows[1:]:
            item = {}
            for idx, header in enumerate(headers):
                item[header] = row[idx] if idx < len(row) else None
            data.append(item)

        return data

    def validate_rows(self, rows):
        report = []
        seen_nip = set()

        for idx, row in enumerate(rows, start=1):
            nip = clean_nip(row.get("NIP"))
            nama = normalize_text(row.get("NAMA"))
            jenis_kepegawaian = normalize_text(row.get("JENIS KEPEGAWAIAN"))
            jabatan_raw = normalize_text(row.get("JABATAN"))
            unit_raw = normalize_text(row.get("UNIT KERJA"))

            status = "OK"
            notes = []

            instansi = None
            jabatan = None
            pegawai = None
            pegawai_action = ""
            user_action = ""

            if not nip or len(nip) != 18:
                status = "INVALID_NIP"
                notes.append("NIP wajib 18 digit")

            if nip and nip in seen_nip:
                status = "DUPLICATE_NIP_IN_FILE"
                notes.append("NIP duplikat dalam file import")

            if nip:
                seen_nip.add(nip)

            if status == "OK" and not nama:
                status = "INVALID_NAMA"
                notes.append("Nama kosong")

            if status == "OK" and not unit_raw:
                status = "INVALID_UNIT"
                notes.append("Unit kerja kosong")

            if status == "OK" and not jabatan_raw:
                status = "INVALID_JABATAN"
                notes.append("Jabatan kosong")

            if status == "OK":
                instansi = RefInstansi.objects.filter(nama__iexact=unit_raw).first()
                if not instansi:
                    status = "UNMATCHED_UNIT"
                    notes.append(f"Unit kerja tidak ditemukan: {unit_raw}")

            if status == "OK":
                jabatan = MstJabatan.objects.filter(
                    nama_jabatan__iexact=jabatan_raw,
                    instansi=instansi,
                ).first()

                if not jabatan:
                    status = "UNMATCHED_JABATAN"
                    notes.append(f"Jabatan tidak ditemukan pada unit kerja tersebut: {jabatan_raw}")

            if status == "OK":
                pegawai_qs = MstDataPegawai.objects.filter(nip_hash=hash_lookup(nip))

                if pegawai_qs.count() > 1:
                    status = "DUPLICATE_NIP_IN_DB"
                    notes.append("Ada lebih dari satu pegawai dengan nip_hash yang sama")
                else:
                    pegawai = pegawai_qs.first()

            if status == "OK":
                if pegawai:
                    pegawai_action = "UPDATE_PEGAWAI"
                    user_action = "LINK_OR_UPDATE_USER"
                else:
                    pegawai_action = "CREATE_PEGAWAI"
                    user_action = "CREATE_USER"

            report.append({
                "row_no": idx,
                "nip": nip,
                "nama": nama,
                "jenis_kepegawaian": jenis_kepegawaian,
                "jabatan": jabatan_raw,
                "unit_kerja": unit_raw,
                "status": status,
                "note": "; ".join(notes),
                "pegawai_id": pegawai.id if pegawai else "",
                "instansi_id": instansi.id if instansi else "",
                "jabatan_id": jabatan.id_jabatan if jabatan else "",
                "pegawai_action": pegawai_action,
                "user_action": user_action,
            })

        return report

    def commit_rows(self, report, create_user=False):
        for r in report:
            nip = r["nip"]
            nama = r["nama"]

            instansi = RefInstansi.objects.get(id=r["instansi_id"])
            jabatan = MstJabatan.objects.get(id_jabatan=r["jabatan_id"])

            pegawai = MstDataPegawai.objects.filter(nip_hash=hash_lookup(nip)).first()

            if pegawai:
                pegawai.nip = nip
                pegawai.nama_lengkap = nama
                pegawai.id_instansi = instansi
                pegawai.id_jabatan = jabatan
                pegawai.is_active = True
                pegawai.updatedby = IMPORT_USER
                pegawai.updateddate = timezone.now()
                pegawai.save(update_fields=[
                    "nip",
                    "nama_lengkap",
                    "nip_hash",
                    "nama_lengkap_hash",
                    "id_instansi",
                    "id_jabatan",
                    "is_active",
                    "updatedby",
                    "updateddate",
                ])
            else:
                pegawai = MstDataPegawai.objects.create(
                    nip=nip,
                    nama_lengkap=nama,
                    id_instansi=instansi,
                    id_jabatan=jabatan,
                    is_active=True,
                    createdby=IMPORT_USER,
                    createddate=timezone.now(),
                )

            if create_user:
                self.create_or_link_user(
                    nip=nip,
                    nama=nama,
                    pegawai=pegawai,
                    instansi=instansi,
                )

    def create_or_link_user(self, nip, nama, pegawai, instansi):
        User = get_user_model()
        username = make_generated_username(nip)
        username_hash = hash_lookup(username)

        user = User.objects.filter(username=username).first()

        if not user:
            user = User.objects.filter(username_hash=username_hash).first()

        expired_at = timezone.now() - timezone.timedelta(days=91)

        if not user:
            user = User(
                username=username,
                username_hash=username_hash,
                pegawai=pegawai,
                instansi=instansi,
                full_name=nama,
                is_active=True,
                must_change_password=True,
                password_changed_at=expired_at,
            )
            user.set_password(nip)
            user.save()
            return user

        changed = []

        if getattr(user, "username_hash", None) != username_hash:
            user.username_hash = username_hash
            changed.append("username_hash")

        if user.pegawai_id != pegawai.id:
            user.pegawai = pegawai
            changed.append("pegawai")

        if user.instansi_id != instansi.id:
            user.instansi = instansi
            changed.append("instansi")

        if not user.full_name:
            user.full_name = nama
            changed.append("full_name")

        if not user.is_active:
            user.is_active = True
            changed.append("is_active")

        if changed:
            user.save(update_fields=changed)

        return user

    def write_report(self, report, report_path):
        fieldnames = [
            "row_no",
            "nip",
            "nama",
            "jenis_kepegawaian",
            "jabatan",
            "unit_kerja",
            "status",
            "note",
            "pegawai_id",
            "instansi_id",
            "jabatan_id",
            "pegawai_action",
            "user_action",
        ]

        with open(report_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(report)